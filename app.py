from concurrent.futures import ThreadPoolExecutor
executor = ThreadPoolExecutor(max_workers=8)

from flask import Flask, render_template, request, send_file, jsonify, session
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from io import BytesIO
import mysql.connector
import threading
import subprocess, os, tempfile
import uuid

app = Flask(__name__)
UPLOAD_LIMIT = 150

# secret key untuk session
app.secret_key = "ganti-ini-ke-random-secret-key"

# ================================
# GLOBAL PROGRESS (tidak dipakai lagi, tapi biarkan saja)
# ================================
progress = {"total": 0, "done": 0}
progress_lock = threading.Lock()

# ================================
# PROGRESS per-user
# ================================
user_progress = {}

@app.before_request
def ensure_uid():
    if "uid" not in session:
        session["uid"] = str(uuid.uuid4())


# ================================
# MYSQL
# ================================
def db():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="db_wilayah"
    )


# ================================
# (DIHAPUS) PDF COMPRESS GHOSTSCRIPT
# ================================
# def compress_pdf_for_processing(file_storage):
#     pass   # FUNGSI DIHAPUS TOTAL


# ================================
# PROGRESS ENDPOINT
# ================================
@app.route("/progress")
def get_progress():
    uid = session.get("uid")
    if not uid or uid not in user_progress:
        return jsonify({"total": 0, "done": 0})
    return jsonify(user_progress[uid])


# ================================
# DATA WILAYAH ENDPOINT
# ================================
@app.route("/get_provinsi")
def get_provinsi():
    con = db()
    cur = con.cursor(dictionary=True)
    cur.execute(
        "SELECT province_code AS code, province AS nama FROM provinces ORDER BY province"
    )
    return jsonify(cur.fetchall())


@app.route("/get_kabkot/<prov_code>")
def get_kabkot(prov_code):
    con = db()
    cur = con.cursor(dictionary=True)
    cur.execute(
        "SELECT city_code AS code, city AS nama FROM cities WHERE province_code=%s ORDER BY city",
        (prov_code,),
    )
    return jsonify(cur.fetchall())


@app.route("/get_kecamatan/<city_code>")
def get_kecamatan(city_code):
    con = db()
    cur = con.cursor(dictionary=True)
    cur.execute(
        "SELECT district_code AS code, district AS nama FROM districts WHERE city_code=%s ORDER BY district",
        (city_code,),
    )
    return jsonify(cur.fetchall())


@app.route("/get_deskel/<district_code>")
def get_deskel(district_code):
    con = db()
    cur = con.cursor(dictionary=True)
    cur.execute(
        "SELECT village_code AS code, village AS nama FROM villages WHERE district_code=%s ORDER BY village",
        (district_code,),
    )
    return jsonify(cur.fetchall())


# ================================
# FORMAT RT/RW
# ================================
def format_rt_rw_columns(df):
    def format_value(value):
        digits = "".join(filter(str.isdigit, str(value)))
        return digits[-3:] if digits else ""
    df["rt"] = df["rt"].apply(format_value)
    df["rw"] = df["rw"].apply(format_value)
    return df


# ================================
# PARSE PDF -> DATAFRAME
# ================================
def convert_pdf_to_dataframe(file_path):

    rows = []

    def valid(row):
        try:
            int(str(row[0]).strip())
            return True
        except:
            return False

    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not any(row):
                        continue
                    if str(row[0]).strip().lower() in ["no", "nomor", "nama"]:
                        continue
                    if not valid(row):
                        continue
                    rows.append(row)

    if not rows:
        return pd.DataFrame(columns=[
            "no","nama","jenis_kelamin","usia",
            "provinsi","kode_provinsi","kabkot","kode_kabkot",
            "kecamatan","kode_kecamatan","deskel","kode_deskel",
            "tps","alamat","rt","rw","ket"
        ])

    max_len = max(len(r) for r in rows)
    for r in rows:
        while len(r) < max_len:
            r.append("")

    df = pd.DataFrame(rows).iloc[:, :8]
    df.columns = ["no","nama","jenis_kelamin","usia","alamat","rt","rw","ket"]

    # hapus baris "1 2 3 4 5 6 7 8"
    if not df.empty:
        f = df.iloc[0].astype(str).str.strip().tolist()
        if f[:8] == [str(i) for i in range(1, 9)]:
            df = df.iloc[1:]

    df = df[df["no"].apply(lambda x: str(x).isdigit())]
    df = df.reset_index(drop=True)

    for col in [
        "provinsi","kode_provinsi","kabkot","kode_kabkot",
        "kecamatan","kode_kecamatan","deskel","kode_deskel","tps"
    ]:
        df[col] = ""

    df = format_rt_rw_columns(df)

    return df.reindex(columns=[
        "no","nama","jenis_kelamin","usia",
        "provinsi","kode_provinsi","kabkot","kode_kabkot",
        "kecamatan","kode_kecamatan","deskel","kode_deskel",
        "tps","alamat","rt","rw","ket"
    ])


# ================================
# MAIN ROUTE
# ================================
@app.route("/", methods=["GET", "POST"])
def upload_file():

    if request.method == "POST":

        prov_code = request.form["provinsi"]
        kab_code  = request.form["kabkot"]
        kec_code  = request.form["kecamatan"]
        des_code  = request.form["deskel"]

        con = db()
        cur = con.cursor(dictionary=True)

        cur.execute(
            "SELECT province AS nama_prov, province_code AS kode_prov FROM provinces WHERE province_code=%s",
            (prov_code,),
        )
        prov = cur.fetchone()

        cur.execute(
            "SELECT city AS nama_kab, city_code AS kode_kab FROM cities WHERE city_code=%s",
            (kab_code,),
        )
        kab = cur.fetchone()

        cur.execute(
            "SELECT district AS nama_kec, district_code AS kode_kec FROM districts WHERE district_code=%s",
            (kec_code,),
        )
        kec = cur.fetchone()

        cur.execute(
            "SELECT village AS nama_des, village_code AS kode_des FROM villages WHERE village_code=%s",
            (des_code,),
        )
        des = cur.fetchone()

        files = request.files.getlist("pdf_file[]")

        if len(files) == 0:
            return "Tidak ada file upload!"
        if len(files) > UPLOAD_LIMIT:
            return "Maksimal 150 file!"

        uid = session["uid"]

        # reset progress user ini
        with progress_lock:
            user_progress[uid] = {"total": len(files), "done": 0}

        # STEP 1: langsung proses PDF tanpa compress
        temp_paths = []
        for f in files:
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
            f.save(temp_file.name)
            temp_paths.append(temp_file.name)

        # STEP 2: parse PDF paralel + update progress
        def parse_with_progress(path):
            df = convert_pdf_to_dataframe(path)
            with progress_lock:
                if uid in user_progress:
                    user_progress[uid]["done"] += 1
            return df

        dfs = list(executor.map(parse_with_progress, temp_paths))

        # hapus file tmp
        for p in temp_paths:
            try: os.remove(p)
            except: pass

        wilayah = des["nama_des"].replace(" ", "_").upper()

        wb = Workbook()
        wb.remove(wb.active)

        for df, f in zip(dfs, files):

            tps_part = f.filename.split("TPS")[-1]
            tps_number = "".join(filter(str.isdigit, tps_part)) or "0"

            sheet = wb.create_sheet(f"TPS {int(tps_number)}")

            df["provinsi"] = prov["nama_prov"]
            df["kode_provinsi"] = prov["kode_prov"]

            df["kabkot"] = kab["nama_kab"]
            df["kode_kabkot"] = kab["kode_kab"]

            df["kecamatan"] = kec["nama_kec"]
            df["kode_kecamatan"] = kec["kode_kec"]

            df["deskel"] = des["nama_des"]
            df["kode_deskel"] = des["kode_des"]

            df["tps"] = tps_number.zfill(3)

            for col_i, col in enumerate(df.columns, start=1):
                sheet.cell(row=1, column=col_i, value=col)

            for row_i, row in df.iterrows():
                for col_i, val in enumerate(row, start=1):
                    sheet.cell(row=row_i + 2, column=col_i, value=val)

        buffer = BytesIO()
        file_name = f"{wilayah}.xlsx"
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=file_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return render_template("index.html")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
